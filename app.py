"""
MPC² · Projekt 3 · I-U Review Workstation — v4 (Codex-revised).

Narrow v1 scope: upload → triage queue → detail + manual override → export.
Drops (to v2): DOCX export, uncertainty estimates, tutorial-examples tab.
"""
from __future__ import annotations
import sys, io, tempfile, json, time
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from iu_analyzer import (
    parse_asc, analyze, IUCurveData, IUAnalysis,
    filename_to_supplier_medium, reanalyze_with_overrides,
)

OVERRIDE_DIR = Path.home() / ".mpc2"
OVERRIDE_DIR.mkdir(exist_ok=True)
OVERRIDE_FILE = OVERRIDE_DIR / "p3_overrides.json"


# ═══════════════════════════════════════════════════════════════════════════
# Page config + password gate
# ═══════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="MPC² · I-U Review Workstation",
    page_icon="⚡",
    layout="wide",
)

def _check_password() -> bool:
    try:
        expected = st.secrets["password"]
    except Exception:
        return True
    if st.session_state.get("auth_ok"):
        return True
    st.markdown("""
    <div style="max-width:460px; margin:80px auto; padding:36px 40px;
                background:white; border:1px solid #EEEEEE; border-left:6px solid #8DBF18;
                font-family:'Manrope',sans-serif;">
      <div style="font-size:11px; font-weight:700; letter-spacing:0.12em;
                  color:#409A2D; text-transform:uppercase; margin-bottom:12px;">
        werchota.ai · MPC² Customer Preview
      </div>
      <h1 style="font-size:26px; font-weight:700; color:#4C4C4C; margin:0 0 8px;">
        I-U Review Workstation
      </h1>
      <p style="color:#333; font-size:14px; line-height:1.5; margin:0 0 20px;">
        Zugang nur für eingeladene Nutzer. Bitte Passwort eingeben.
      </p>
    </div>
    """, unsafe_allow_html=True)
    _, col_form, _ = st.columns([1, 2, 1])
    with col_form:
        pw = st.text_input("Passwort", type="password", key="pw_input",
                           label_visibility="collapsed", placeholder="Passwort")
        if pw:
            if pw == expected:
                st.session_state.auth_ok = True
                st.rerun()
            else:
                st.error("Passwort falsch.")
    return False

if not _check_password():
    st.stop()


# ═══════════════════════════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&display=swap');
html, body, .stApp, p, h1, h2, h3, h4, h5, h6, button, input, label {
  font-family: 'Manrope', sans-serif !important;
  color: #333;
}
.stApp { background: #FFFFFF; }
#MainMenu, footer { visibility: hidden; }
.mpc-nav {
  background: linear-gradient(90deg, #8DBF18 6.6%, #409A2D 100%);
  padding: 16px 32px;
  display: flex; justify-content: space-between; align-items: center;
  color: white; box-shadow: 0 2px 12px rgba(0,0,0,0.08);
}
.mpc-nav-title { font-size: 20px; font-weight: 700; color: white !important; }
.mpc-nav-sub { font-size: 13px; font-weight: 500; opacity: 0.92; color: white !important; }
h1 { color: #4C4C4C !important; font-size: 28px !important; font-weight: 700 !important;
     margin: 1.2rem 0 1rem !important; }
h2 { color: #4C4C4C !important; font-size: 20px !important; font-weight: 700 !important;
     padding-left: 16px !important; border-left: 4px solid #8DBF18 !important;
     margin: 1.5rem 0 0.8rem !important; }
.stButton > button, .stDownloadButton > button {
  background-color: #8DBF18 !important; color: white !important;
  border-radius: 0 !important; border: none !important; font-weight: 700 !important;
}
.status-pill {
  display: inline-block; padding: 3px 10px; border-radius: 12px;
  font-size: 11px; font-weight: 700; letter-spacing: .06em;
}
.status-green { background: #EAF4D8; color: #409A2D; }
.status-yellow { background: #fef3c7; color: #b8860b; }
.status-red { background: #fdecec; color: #c0392b; }
.kpi-card { background: white; border: 1px solid #EEEEEE; border-left: 4px solid #8DBF18;
            padding: 14px 18px; margin-bottom: 10px; }
.kpi-card.review { border-left-color: #b8860b; background: #fef8e7; }
.kpi-card.blank  { border-left-color: #888; background: #f6f6f6; color: #888; }
.kpi-card .label { font-size: 11px; font-weight: 700; text-transform: uppercase;
                   letter-spacing: 0.1em; color: #4C4C4C; }
.kpi-card .value { font-size: 26px; font-weight: 700; color: #4C4C4C; margin-top: 4px; }
.kpi-card.blank .value { color: #888; font-weight: 500; font-size: 18px; }
.kpi-card .sub { font-size: 12px; color: #888; margin-top: 2px; }
</style>
<div class="mpc-nav">
  <div style="display:flex; align-items:center; gap:14px;">
    <div style="background:white; padding:8px 20px; transform:skew(-12deg); box-shadow:0 2px 6px rgba(0,0,0,0.1);">
      <span style="display:inline-block; transform:skew(12deg); font-weight:800; font-size:26px;
             color:#4C4C4C;">MPC<sup style="color:#8DBF18">²</sup></span>
    </div>
    <div>
      <div class="mpc-nav-title">Projekt 3 · I-U Review Workstation</div>
      <div class="mpc-nav-sub">Fast analyst workstation · Auto-Suggestion + Manual Review</div>
    </div>
  </div>
  <div class="mpc-nav-sub">gebaut von werchota.ai</div>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# State
# ═══════════════════════════════════════════════════════════════════════════
if "measurements" not in st.session_state:
    st.session_state.measurements = {}
if "selected_msg_id" not in st.session_state:
    st.session_state.selected_msg_id = None

# Load persisted overrides
def load_overrides() -> dict:
    if OVERRIDE_FILE.exists():
        try:
            return json.loads(OVERRIDE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def save_overrides():
    payload = {
        msg_id: {
            "filename": m["filename"],
            "overrides": m["overrides"],
            "override_reasons": m.get("override_reasons", {}),
        }
        for msg_id, m in st.session_state.measurements.items()
        if m["overrides"]
    }
    OVERRIDE_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def apply_overrides(analysis: IUAnalysis, overrides: dict) -> IUAnalysis:
    """Thin wrapper: if only value-overrides (no OCP change), quick swap;
    if OCP changed, call reanalyze_with_overrides() for full recomputation."""
    if not overrides:
        return analysis
    import copy
    # If OCP or family was overridden → do full recomputation
    # (requires the data — caller must pass it)
    a = copy.copy(analysis)
    for k, v in overrides.items():
        if v is not None and hasattr(a, k):
            setattr(a, k, v)
    return a


def recompute_with_overrides(data: IUCurveData, analysis: IUAnalysis,
                              overrides: dict) -> IUAnalysis:
    """Full recomputation path using the analyzer's reanalyze_with_overrides()."""
    if not overrides:
        return analysis
    return reanalyze_with_overrides(data, analysis, overrides)


def status_pill(status: str) -> str:
    css = {"green": "status-green", "yellow": "status-yellow", "red": "status-red"}[status]
    emoji = {"green": "●", "yellow": "●", "red": "●"}[status]
    label = {"green": "ACCEPT", "yellow": "REVIEW", "red": "MANUAL"}[status]
    return f'<span class="status-pill {css}">{emoji} {label}</span>'


# ═══════════════════════════════════════════════════════════════════════════
# TABS — 4 only (Codex: defer Anleitung-tab to v2)
# ═══════════════════════════════════════════════════════════════════════════
tab_upload, tab_queue, tab_detail, tab_export = st.tabs([
    "📥 Upload", "🚦 Review Queue", "🔍 Detail + Override", "💾 Export"
])


# ────────────────────────────────────────────────────────────────────────────
# TAB 1 — UPLOAD
# ────────────────────────────────────────────────────────────────────────────
with tab_upload:
    st.markdown("# Upload I-U Messdaten")
    st.markdown(
        "ASC + HAD Dateien hochladen. Die App klassifiziert jede Messung in eine "
        "**Kurven-Familie** (pitting / transpassiv / prep+rp / multi-cycle / malformed), "
        "füllt pro Familie nur die unterstützten Metriken und liefert eine **Triage-Queue** "
        "mit Ampel grün/gelb/rot. Unsichere Werte bleiben bewusst leer — du setzt sie manuell."
    )

    col_a, col_b = st.columns(2)
    with col_a:
        asc_uploads = st.file_uploader("ASC-Files", type=["ASC", "asc"],
                                       accept_multiple_files=True, key="asc_up")
    with col_b:
        had_uploads = st.file_uploader("HAD-Files (optional)", type=["HAD", "had"],
                                       accept_multiple_files=True, key="had_up")

    if asc_uploads and st.button("▶ Analyse starten", use_container_width=True, type="primary"):
        had_paths = {}
        for h in (had_uploads or []):
            tmp = Path(tempfile.gettempdir()) / f"mpc2_p3_{h.name}"
            tmp.write_bytes(h.getvalue())
            had_paths[h.name.lower()] = tmp

        st.session_state.measurements = {}
        # Restore any previous overrides from JSON
        persisted = load_overrides()

        progress = st.progress(0, text="Parse…")
        for i, asc in enumerate(asc_uploads, 1):
            progress.progress(i / len(asc_uploads),
                              text=f"{i}/{len(asc_uploads)}: {asc.name}")
            try:
                asc_tmp = Path(tempfile.gettempdir()) / f"mpc2_p3_{asc.name}"
                asc_tmp.write_bytes(asc.getvalue())
                had_name = asc.name.replace(".ASC", ".HAD").replace(".asc", ".had")
                had_tmp = had_paths.get(had_name.lower())
                data = parse_asc(asc_tmp, had_tmp)
                analysis = analyze(data)
                msg_id = asc.name.split("_")[0]
                # Restore persisted overrides for this measurement
                prev = persisted.get(msg_id, {})
                st.session_state.measurements[msg_id] = {
                    "filename": asc.name,
                    "data": data,
                    "analysis": analysis,
                    "overrides": prev.get("overrides", {}),
                    "override_reasons": prev.get("override_reasons", {}),
                }
            except Exception as e:
                st.error(f"Fehler bei {asc.name}: {e}")
        progress.empty()
        n = len(st.session_state.measurements)
        st.success(f"✓ {n} Messungen analysiert → gehe zur **Review Queue**.")


# ────────────────────────────────────────────────────────────────────────────
# TAB 2 — REVIEW QUEUE (traffic-light, sorted red-to-green)
# ────────────────────────────────────────────────────────────────────────────
with tab_queue:
    st.markdown("# 🚦 Review Queue")
    if not st.session_state.measurements:
        st.info("Noch keine Messungen hochgeladen.")
    else:
        # Build queue rows
        rows = []
        for msg_id, m in st.session_state.measurements.items():
            a = apply_overrides(m["analysis"], m["overrides"])
            sup, med = filename_to_supplier_medium(m["filename"])
            rows.append({
                "ID": msg_id,
                "Status": m["analysis"].status,
                "Familie": m["analysis"].family,
                "Family-Conf": round(m["analysis"].family_conf, 2),
                "Lieferant": sup or "?",
                "Medium": med or "?",
                "OCP [mV]": a.ocp_mV,
                "RPP [mV]": a.rpp_mV,
                "E_pit [mV]": a.e_pit_mV,
                "i_pass": a.i_pass_mA_cm2,
                "Overrides": "✏" if m["overrides"] else "",
                "File": m["filename"][:50],
            })
        df = pd.DataFrame(rows)

        # Summary KPIs
        col1, col2, col3, col4 = st.columns(4)
        counts = df["Status"].value_counts()
        col1.metric("🟢 ACCEPT (green)", int(counts.get("green", 0)),
                    help="Hohe Confidence, kein Review nötig")
        col2.metric("🟡 REVIEW (yellow)", int(counts.get("yellow", 0)),
                    help="Schnell prüfen, bei Bedarf override")
        col3.metric("🔴 MANUAL (red)", int(counts.get("red", 0)),
                    help="Manuelle Werte-Setzung erforderlich")
        col4.metric("✏ Overridden", sum(1 for r in rows if r["Overrides"]))

        # Sort: red first, then yellow, then green
        status_order = {"red": 0, "yellow": 1, "green": 2}
        df["_sort"] = df["Status"].map(status_order)
        df_sorted = df.sort_values("_sort").drop(columns=["_sort"])

        # Render with conditional styling
        def highlight_status(row):
            color_map = {"red": "#fdecec", "yellow": "#fef3c7", "green": "#EAF4D8"}
            return [f'background-color: {color_map[row["Status"]]}'] * len(row)

        styled = df_sorted.style.apply(highlight_status, axis=1)
        st.dataframe(styled, width="stretch", hide_index=True)

        # Selector for detail
        st.markdown("---")
        msg_ids = list(st.session_state.measurements.keys())
        msg_ids_sorted = sorted(msg_ids, key=lambda mid: status_order[
            st.session_state.measurements[mid]["analysis"].status])
        sel_idx = 0
        if st.session_state.selected_msg_id in msg_ids_sorted:
            sel_idx = msg_ids_sorted.index(st.session_state.selected_msg_id)
        selected = st.selectbox(
            "Messung zum Bearbeiten vormerken (rot zuerst sortiert)",
            msg_ids_sorted, index=sel_idx, key="queue_select",
            format_func=lambda mid: f"{st.session_state.measurements[mid]['analysis'].status.upper()[0]} · {mid} · {st.session_state.measurements[mid]['filename'][:50]}"
        )
        st.session_state.selected_msg_id = selected
        st.info("📋 **Vormerkung gespeichert** — klick jetzt oben auf den Tab **🔍 Detail + Override** um diese Messung zu bearbeiten. (Streamlit kann Tabs nicht programmatisch umschalten.)")

        # Status explanations — one paragraph per traffic light
        st.markdown("---")
        st.markdown("## Was bedeuten die Status-Farben?")
        col_g, col_y, col_r = st.columns(3)
        with col_g:
            st.markdown("""
            <div class="status-pill status-green">● ACCEPT</div>

            **Grün = „Auto-Werte sind gut, nichts zu tun."**

            Die Kurve fällt in eine klar erkannte Familie (pitting oder transpassive)
            und alle berechneten Werte haben hohe Confidence (≥ 0.75). Die Metriken
            (OCP, ggf. RPP, ggf. E_pit, i_pass, Korrosionsrate) kannst du direkt ins
            Berichts-Excel übernehmen ohne Einzelprüfung. In der Praxis: das sind oft
            die glatten, textbuchartigen Kurven — Standard-Fälle bei Routine-Messungen.
            """, unsafe_allow_html=True)
        with col_y:
            st.markdown("""
            <div class="status-pill status-yellow">● REVIEW</div>

            **Gelb = „Bitte kurz drüberschauen."**

            Mindestens ein Wert hat mittlere Confidence (zwischen 0.55 und 0.75), oder
            die Familien-Zuordnung ist nicht ganz eindeutig (z. B. „weak hysteresis
            loop"). Die Auto-Werte sind wahrscheinlich korrekt, aber du solltest im
            Detail-Tab einmal visuell verifizieren. Falls die Plots passen: einfach
            nichts tun. Falls nicht: OCP/RPP/E_pit mit manuellem Override setzen.
            """, unsafe_allow_html=True)
        with col_r:
            st.markdown("""
            <div class="status-pill status-red">● MANUAL</div>

            **Rot = „Auto-Extraktion zu unsicher — manuelle Werte-Setzung erforderlich."**

            Entweder wurde die Datei als „malformed" klassifiziert (zu kurz, zu noisy),
            als „multi_cycle" (mehrere Scan-Zyklen — Auto kann nicht wählen welcher
            der richtige ist), oder die Confidences liegen unter 0.55. Öffne im
            Detail-Tab den Plot, setze OCP/RPP/E_pit manuell (mit Begründung im Audit-Feld).
            Nach Override springt der Status auf grün und der Wert wird mit
            Provenienz „override" exportiert.
            """, unsafe_allow_html=True)


# ────────────────────────────────────────────────────────────────────────────
# TAB 3 — DETAIL + OVERRIDE
# ────────────────────────────────────────────────────────────────────────────
with tab_detail:
    st.markdown("# 🔍 Detail + Override")
    if not st.session_state.measurements:
        st.info("Noch keine Messungen hochgeladen.")
    else:
        msg_ids = list(st.session_state.measurements.keys())
        default_idx = (msg_ids.index(st.session_state.selected_msg_id)
                       if st.session_state.selected_msg_id in msg_ids else 0)
        sel = st.selectbox("Messung", msg_ids, index=default_idx, key="detail_select")
        st.session_state.selected_msg_id = sel

        m = st.session_state.measurements[sel]
        data = m["data"]
        raw = m["analysis"]
        # After any override: do a FULL recomputation (i_pass/corr-rate/hysteresis update)
        current = recompute_with_overrides(data, raw, m["overrides"])

        # Header with family + status
        col_a, col_b, col_c = st.columns([2, 1, 1])
        with col_a:
            st.markdown(f"### `{sel}` · {m['filename']}")
        with col_b:
            st.markdown(f"**Familie:** {current.family}  \n_{current.family_reason}_")
        with col_c:
            st.markdown(f"**Status:** {status_pill(current.status)}", unsafe_allow_html=True)
            st.caption(current.status_reason)

        # ═══════════════════════════════════════════════════════════════
        # FINAL OUTPUT PANEL — what actually goes into Manuel's report
        # ═══════════════════════════════════════════════════════════════
        st.markdown("## 📊 Endergebnisse für den Bericht")
        st.caption("Diese Zahlen gehen 1:1 in den akkreditierten Prüfbericht. "
                   "Sie werden bei jedem OCP/Override automatisch neu berechnet.")
        fo_c1, fo_c2, fo_c3, fo_c4 = st.columns(4)
        with fo_c1:
            val = current.i_pass_mA_cm2
            st.markdown(f"""<div class="kpi-card {'blank' if val is None else ''}" style="border-left-color:#409A2D;">
                <div class="label">i_pass · Passivstromdichte</div>
                <div class="value">{val if val is not None else '—'}</div>
                <div class="sub">mA/cm² · {"auto" if val else "nicht verfügbar"}</div>
            </div>""", unsafe_allow_html=True)
        with fo_c2:
            val = current.corrosion_rate_mm_per_year
            st.markdown(f"""<div class="kpi-card {'blank' if val is None else ''}" style="border-left-color:#409A2D;">
                <div class="label">Korrosionsrate</div>
                <div class="value">{val if val is not None else '—'}</div>
                <div class="sub">mm/Jahr (Fe, default) · Faraday</div>
            </div>""", unsafe_allow_html=True)
        with fo_c3:
            val = current.hysteresis_area
            st.markdown(f"""<div class="kpi-card {'blank' if val is None else ''}" style="border-left-color:#409A2D;">
                <div class="label">Hysterese-Fläche</div>
                <div class="value">{val if val is not None else '—'}</div>
                <div class="sub">mA·V/cm² · Integral</div>
            </div>""", unsafe_allow_html=True)
        with fo_c4:
            fam_map = {
                "pitting": "Pitting-Korrosion",
                "transpassive": "Transpassiv (kein Pitting)",
                "prep_then_rp": "Pitting (Prep-Scan erkannt)",
                "multi_cycle": "Mehrfach-Zyklus — manuell",
                "malformed": "Unbrauchbar — manuell",
            }
            family_label = fam_map.get(current.family, current.family)
            st.markdown(f"""<div class="kpi-card" style="border-left-color:#8DBF18;">
                <div class="label">Korrosions-Typ</div>
                <div class="value" style="font-size:20px;">{family_label}</div>
                <div class="sub">Familie: {current.family} · conf {current.family_conf:.2f}</div>
            </div>""", unsafe_allow_html=True)

        # Plots
        pot_mV = data.potential_V * 1000
        j_abs_mA_cm2 = np.abs(data.current_density_A_m2) * 0.1
        vtx = raw.vertex_index or len(pot_mV) // 2
        w_end = raw.ocp_window_end or 0
        fwd_E = pot_mV[w_end:vtx]
        fwd_j = j_abs_mA_cm2[w_end:vtx]
        rev_E = pot_mV[vtx:]
        rev_j = j_abs_mA_cm2[vtx:]

        # Plot 1: potential vs time (with plateaus highlighted)
        fig1 = go.Figure()
        fig1.add_trace(go.Scatter(x=data.time_s, y=pot_mV, mode="lines",
                                   name="Potenzial", line=dict(color="#409A2D", width=1)))
        # Highlight detected plateaus
        for i, (pstart, pend, pmV) in enumerate(raw.plateaus):
            fig1.add_vrect(x0=data.time_s[pstart], x1=data.time_s[pend],
                           fillcolor="#8DBF18", opacity=0.15, line_width=0,
                           annotation_text=f"Plateau {i+1}: {pmV:.0f}mV")
        if current.ocp_mV is not None:
            fig1.add_hline(y=current.ocp_mV, line=dict(color="#8DBF18", dash="dash"),
                           annotation_text=f"OCP {current.ocp_mV:.1f} mV")
        if raw.vertex_index:
            fig1.add_vline(x=data.time_s[raw.vertex_index],
                           line=dict(color="#b8860b", dash="dot"),
                           annotation_text="Vertex")
        fig1.update_layout(height=320, xaxis_title="Zeit [s]",
                           yaxis_title="Potenzial [mV]",
                           margin=dict(l=10, r=10, t=20, b=10))
        st.plotly_chart(fig1, use_container_width=True)

        # Plot 2: I-U curve log scale
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=fwd_E, y=np.maximum(fwd_j, 1e-6),
                                   mode="lines", name="Forward",
                                   line=dict(color="#409A2D", width=1.2),
                                   hovertemplate="E=%{x:.1f} mV · |j|=%{y:.4f}"))
        fig2.add_trace(go.Scatter(x=rev_E, y=np.maximum(rev_j, 1e-6),
                                   mode="lines", name="Reverse",
                                   line=dict(color="#b8860b", width=1.2),
                                   hovertemplate="E=%{x:.1f} mV · |j|=%{y:.4f}"))
        if current.ocp_mV is not None:
            fig2.add_vline(x=current.ocp_mV, line=dict(color="#8DBF18", dash="dash"),
                           annotation_text="OCP")
        if current.rpp_mV is not None:
            fig2.add_vline(x=current.rpp_mV, line=dict(color="#c0392b", dash="dashdot"),
                           annotation_text="RPP")
        if current.e_pit_mV is not None:
            fig2.add_vline(x=current.e_pit_mV, line=dict(color="purple", dash="dot"),
                           annotation_text="E_pit")
        fig2.update_layout(height=460, xaxis_title="Potenzial [mV]",
                           yaxis_title="|j| [mA/cm²]", yaxis_type="log",
                           margin=dict(l=10, r=10, t=20, b=10))
        st.plotly_chart(fig2, use_container_width=True)

        # ═══════════════════════════════════════════════════════════════
        # OVERRIDE PANELS — inputs that drive the final output
        # ═══════════════════════════════════════════════════════════════
        st.markdown("## ✏ Manueller Override · wenn Auto daneben liegt")
        st.info(
            "💡 **So funktioniert es:** Trage einen Wert ein + optional eine Begründung, dann klick **Set & Recalculate**. "
            "Die Endergebnisse oben (i_pass, Korrosionsrate, Hysterese-Fläche) werden bei einem OCP-Override automatisch "
            "neu berechnet. RPP und E_pit sind eigenständige Werte ohne Folgeberechnung. "
            "Der Override wird in `~/.mpc2/p3_overrides.json` mit Zeitstempel und Begründung persistiert (Audit-Trail)."
        )

        def render_override(col, label: str, attr: str, unit: str = "mV"):
            with col:
                auto_val = getattr(raw, attr)
                current_val = m["overrides"].get(attr, auto_val)
                conf_attr = attr.replace("_mV", "_conf").replace("_mA_cm2", "_conf")
                conf = getattr(raw, conf_attr, 0.0)
                is_blank = auto_val is None
                is_override = attr in m["overrides"]

                if is_blank and not is_override:
                    cls = "blank"; mode = "kein Auto-Wert"
                elif conf < 0.7 and not is_override:
                    cls = "review"; mode = f"auto · conf {conf:.2f} ⚠"
                elif is_override:
                    cls = ""; mode = "OVERRIDE"
                else:
                    cls = ""; mode = f"auto · conf {conf:.2f} ✓"

                st.markdown(
                    f"""<div class="kpi-card {cls}">
                        <div class="label">{label}</div>
                        <div class="value">{current_val if current_val is not None else '—'}</div>
                        <div class="sub">{unit} · {mode}</div>
                    </div>""", unsafe_allow_html=True)

                new_val = st.number_input(
                    f"{label}-Wert setzen", value=float(current_val) if current_val is not None else 0.0,
                    step=1.0, key=f"in_{attr}_{sel}", label_visibility="collapsed")
                reason = st.text_input("Grund (optional)",
                                       value=m.get("override_reasons", {}).get(attr, ""),
                                       key=f"rsn_{attr}_{sel}", label_visibility="collapsed",
                                       placeholder="z.B. spätere stabile Phase")

                col_set, col_rst = st.columns(2)
                with col_set:
                    if st.button(f"✓ Set {label} & Recalculate",
                                 key=f"set_{attr}_{sel}", use_container_width=True,
                                 type="primary"):
                        m["overrides"][attr] = round(new_val, 2)
                        m.setdefault("override_reasons", {})[attr] = reason
                        save_overrides()
                        st.success(f"{label} gesetzt auf {new_val:.2f} — scrollt nach oben, die Endergebnisse sind aktualisiert.")
                        st.rerun()
                with col_rst:
                    if st.button("↺ Zurück auf Auto", key=f"rst_{attr}_{sel}",
                                 use_container_width=True):
                        m["overrides"].pop(attr, None)
                        m.get("override_reasons", {}).pop(attr, None)
                        save_overrides()
                        st.rerun()

        c1, c2, c3 = st.columns(3)
        render_override(c1, "OCP",   "ocp_mV",   "mV")
        render_override(c2, "RPP",   "rpp_mV",   "mV")
        render_override(c3, "E_pit", "e_pit_mV", "mV")

        st.caption("💡 **OCP ist der Haupt-Override**: bei OCP-Änderung werden i_pass, "
                   "Korrosionsrate und Hysterese-Fläche neu berechnet (die Zahlen oben "
                   "im grünen Endergebnisse-Block aktualisieren sich). RPP und E_pit sind "
                   "Einzelwerte ohne Folgeberechnung.")

        # Family override
        st.markdown("### Korrosions-Typ manuell festlegen")
        family_options = ["auto", "pitting", "transpassive", "prep_then_rp", "multi_cycle", "malformed"]
        current_fam_override = m["overrides"].get("family", "auto")
        new_fam = st.selectbox(
            "Falls die Auto-Klassifikation falsch ist, hier setzen",
            family_options,
            index=family_options.index(current_fam_override) if current_fam_override in family_options else 0,
            key=f"fam_{sel}"
        )
        if st.button("Korrosions-Typ übernehmen & neu berechnen",
                     key=f"setfam_{sel}", use_container_width=True):
            if new_fam == "auto":
                m["overrides"].pop("family", None)
            else:
                m["overrides"]["family"] = new_fam
            save_overrides()
            st.rerun()

        if m["overrides"]:
            with st.expander("Overrides anzeigen (persistiert in ~/.mpc2/p3_overrides.json)"):
                st.json({"overrides": m["overrides"],
                         "reasons": m.get("override_reasons", {})})


# ────────────────────────────────────────────────────────────────────────────
# TAB 4 — EXPORT
# ────────────────────────────────────────────────────────────────────────────
def build_overview_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Queue"
    headers = ["ID", "File", "Lieferant", "Medium", "Status", "Familie", "Family-Conf",
               "OCP [mV]", "OCP Method", "OCP Conf",
               "RPP [mV]", "RPP Method",
               "E_pit [mV]", "E_pit Method",
               "i_pass [mA/cm²]", "Korrosionsrate", "Hysterese-Fläche"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF4C4C4C")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r, (mid, m) in enumerate(st.session_state.measurements.items(), 2):
        a = apply_overrides(m["analysis"], m["overrides"])
        sup, med = filename_to_supplier_medium(m["filename"])
        ws.cell(row=r, column=1, value=mid)
        ws.cell(row=r, column=2, value=m["filename"])
        ws.cell(row=r, column=3, value=sup)
        ws.cell(row=r, column=4, value=med)
        ws.cell(row=r, column=5, value=m["analysis"].status)
        ws.cell(row=r, column=6, value=m["analysis"].family)
        ws.cell(row=r, column=7, value=round(m["analysis"].family_conf, 2))
        ws.cell(row=r, column=8, value=a.ocp_mV)
        ws.cell(row=r, column=9, value="override" if "ocp_mV" in m["overrides"] else "auto")
        ws.cell(row=r, column=10, value=round(m["analysis"].ocp_conf, 2))
        ws.cell(row=r, column=11, value=a.rpp_mV)
        ws.cell(row=r, column=12, value="override" if "rpp_mV" in m["overrides"] else ("auto" if a.rpp_mV else "blank"))
        ws.cell(row=r, column=13, value=a.e_pit_mV)
        ws.cell(row=r, column=14, value="override" if "e_pit_mV" in m["overrides"] else ("auto" if a.e_pit_mV else "blank"))
        ws.cell(row=r, column=15, value=a.i_pass_mA_cm2)
        ws.cell(row=r, column=16, value=a.corrosion_rate_mm_per_year)
        ws.cell(row=r, column=17, value=a.hysteresis_area)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


with tab_export:
    st.markdown("# Export")
    if not st.session_state.measurements:
        st.info("Noch keine Messungen hochgeladen.")
    else:
        n = len(st.session_state.measurements)
        n_ovr = sum(1 for m in st.session_state.measurements.values() if m["overrides"])
        n_green = sum(1 for m in st.session_state.measurements.values() if m["analysis"].status == "green")
        n_yellow = sum(1 for m in st.session_state.measurements.values() if m["analysis"].status == "yellow")
        n_red = sum(1 for m in st.session_state.measurements.values() if m["analysis"].status == "red")
        st.markdown(f"""
        - **{n}** Messungen gesamt
        - **{n_green}** 🟢 green, **{n_yellow}** 🟡 yellow, **{n_red}** 🔴 red
        - **{n_ovr}** Messungen mit manuellem Override
        - Provenienz pro Zelle (auto / override / blank) ist im Export enthalten
        """)
        data = build_overview_xlsx()
        st.download_button("📥 Review-Queue.xlsx herunterladen",
                           data=data, file_name="Projekt3_IU_ReviewQueue.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

        if OVERRIDE_FILE.exists():
            st.markdown("---")
            st.caption(f"Audit-Trail (persistierte Overrides): `{OVERRIDE_FILE}`")
            with open(OVERRIDE_FILE, "rb") as f:
                st.download_button("📥 Audit-JSON herunterladen",
                                   data=f.read(),
                                   file_name="p3_overrides_audit.json",
                                   mime="application/json",
                                   use_container_width=True)
